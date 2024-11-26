from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import unicodedata
from openpyxl import Workbook, load_workbook
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def normalizar_texto(texto):
    """
    Remove inconsistências de encoding e normaliza acentos.
    """
    return unicodedata.normalize('NFC', texto)

def obter_valores_frete_selenium(url, cep):
    """
    Calcula frete para um único CEP usando Selenium.
    """
    driver = webdriver.Chrome()
    driver.get(url)

    try:
        wait = WebDriverWait(driver, 10)
        campo_cep = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'x-product__shipping-input')))
        
        campo_cep.clear()
        campo_cep.send_keys(cep)
        
        botao_calcular = driver.find_element(By.CLASS_NAME, 'x-product__shipping-submit')
        botao_calcular.click()
        
        time.sleep(6) 
        
        linhas_frete = driver.find_elements(By.CLASS_NAME, 'x-product__shipping-row')
        opcoes_frete = []
        for linha in linhas_frete:
            valor = linha.find_element(By.CLASS_NAME, 'x-product__shipping-value').text
            tipo = linha.find_element(By.CLASS_NAME, 'x-product__shipping-type').text
            prazo = linha.find_element(By.CLASS_NAME, 'x-product__shipping-sla').text
            
            opcoes_frete.append({
                'valor': valor.strip(),
                'tipo': tipo.strip(),
                'prazo': prazo.strip()
            })
        
        return opcoes_frete
    
    finally:
        driver.quit()

def calcular_frete_para_ceps(url, lista_ceps):
    """
    Calcula frete para múltiplos CEPs.
    """
    resultados = {}
    for cep in lista_ceps:
        print(f"\nCalculando frete para o CEP: {cep}...")
        opcoes_frete = obter_valores_frete_selenium(url, cep)
        resultados[cep] = opcoes_frete

    return resultados

def salvar_em_excel(caminho_arquivo, dados):
    """
    Salva os dados no Excel.
    """
    try:
        try:
            workbook = load_workbook(caminho_arquivo)
        except FileNotFoundError:
            workbook = Workbook()
        
        if 'Fretes' in workbook.sheetnames:
            sheet = workbook['Fretes']
        else:
            sheet = workbook.create_sheet('Fretes')

        # Cabeçalhos
        if sheet.max_row == 1 and sheet.cell(1, 1).value is None:
            sheet.append(['CEP', 'Tipo', 'Valor', 'Prazo'])

        # Adicionar os dados
        for cep, opcoes in dados.items():
            for opcao in opcoes:
                sheet.append([cep, opcao['tipo'], opcao['valor'], opcao['prazo']])
        
        # Salvar o arquivo
        workbook.save(caminho_arquivo)
        print(f"Dados salvos em: {caminho_arquivo}")
    except Exception as e:
        print(f"Erro ao salvar no Excel: {e}")


if __name__ == "__main__":
    url_produto = 'https://www.sestini.com.br/mala-de-viagem-bordo-com-rodas-360-gama-3-preto/p'
    lista_ceps = [
    "01000-000",
    "20000-000",
    #"20000-000",
    #"30100-000",
    #"90000-000",
    #80000-000",
    #"40000-000",
    #"60000-000",
    #"70000-000",
    #"50000-000",
    #"69000-000",
    #"66000-000",
    #"74000-000",
    #"29000-000",
    #"88000-000",
    #"64000-000",
    #"65000-000",
    #"59000-000",
    #"57000-000",
    #"49000-000",
    #"79000-000",
    #"78000-000",
    #"77000-000",
    #"69300-000",
    #"68900-000"
    ]

    resultados_frete = calcular_frete_para_ceps(url_produto, lista_ceps)
    
    for cep, opcoes in resultados_frete.items():
        print(f"\nResultados para CEP {cep}:")
        for opcao in opcoes:
            print(f"Tipo: {opcao['tipo']}")
            print(f"Valor: {opcao['valor']}")
            print(f"Prazo: {opcao['prazo']}")
            print("---")

    caminho_excel = r"C:\Users\BG-PROVISORIO\Desktop\Testando códigos do Claude\CEP.xlsx"
    salvar_em_excel(caminho_excel, resultados_frete)